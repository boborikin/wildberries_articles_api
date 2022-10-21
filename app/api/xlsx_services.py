from openpyxl import load_workbook, Workbook
from typing import Generator


def get_articles_from_xslx_file(file: bytes | str) -> list:
    wb = _get_workbook(file)
    wb_values = _get_values_from_workbook(wb)
    return _get_workbook_first_column_values(wb_values)


def _get_workbook(file: bytes | str) -> Workbook:
    return load_workbook(file)


def _get_values_from_workbook(wb: Workbook) -> Generator:
    return wb[wb.sheetnames[0]].values


def _get_workbook_first_column_values(wb_values: Generator) -> list:
    return [value[0] for value in wb_values]
